import importlib

import logging
import os
import shutil
import sys
import scipy.sparse as sparse
import random

import numpy as np
import torch
import torch_geometric.nn as gnn
from torch_geometric.data import Data
import torch.nn.init as init

import itertools
from openpyxl import load_workbook
import pandas as pd

def save_checkpoint(state, is_best, checkpoint_dir, logger=None):
    """Saves model and training parameters at '{checkpoint_dir}/last_checkpoint.pytorch'.
    If is_best==True saves '{checkpoint_dir}/best_checkpoint.pytorch' as well.

    Args:
        state (dict): contains model's state_dict, optimizer's state_dict, epoch
            and best evaluation metric value so far
        is_best (bool): if True state contains the best model seen so far
        checkpoint_dir (string): directory where the checkpoint are to be saved
    """
    def log_info(message):
        if logger is not None:
            logger.info(message)

    if not os.path.exists(checkpoint_dir):
        log_info(
            f"Checkpoint directory does not exists. Creating {checkpoint_dir}")
        os.makedirs(checkpoint_dir)

    #last_file_path = os.path.join(checkpoint_dir, 'last_checkpoint_iter'+str(state['num_iterations'])+'.pytorch')
    #last_file_path = os.path.join(checkpoint_dir, 'last_checkpoint.pytorch')
    #log_info(f"Saving last checkpoint to '{last_file_path}'")
    #torch.save(state, last_file_path)
    if is_best:
        best_file_path = os.path.join(checkpoint_dir, 'best_checkpoint.pytorch')
        log_info(f"Saving best checkpoint to '{best_file_path}'")
        #shutil.copyfile(last_file_path, best_file_path)
        torch.save(state, best_file_path)


def load_checkpoint(checkpoint_path, model, optimizer=None):
    """Loads model and training parameters from a given checkpoint_path
    If optimizer is provided, loads optimizer's state_dict of as well.

    Args:
        checkpoint_path (string): path to the checkpoint to be loaded
        model (torch.nn.Module): model into which the parameters are to be copied
        optimizer (torch.optim.Optimizer) optional: optimizer instance into
            which the parameters are to be copied

    Returns:
        state
    """
    if not os.path.exists(checkpoint_path):
        raise IOError(f"Checkpoint '{checkpoint_path}' does not exist")

    state = torch.load(checkpoint_path)
    #if torch.cuda.device_count() > 1:
    #    model.module.load_state_dict(state['model_state_dict'])
    #else:
    model.load_state_dict(state['model_state_dict'])
    if optimizer is not None:
        optimizer.load_state_dict(state['optimizer_state_dict'])

    return state

def save_model(net, path, filename):
    """Save trained model."""
    if not os.path.exists(path):
        os.makedirs(path)
    torch.save(net.state_dict(),
               os.path.join(path, filename))
    print("save model to: {}".format(os.path.join(path, filename)))

def get_logger(name, level=logging.INFO):
    logger = logging.getLogger(name)
    logger.setLevel(level)
    # Logging to console
    stream_handler = logging.StreamHandler(sys.stdout)
    formatter = logging.Formatter(
        '%(asctime)s [%(threadName)s] %(levelname)s %(name)s - %(message)s')
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    return logger

def get_number_of_learnable_parameters(model):
    if isinstance(model, list):
        num_parameters = 0
        for x_model in model:
            model_parameters = filter(lambda p: p.requires_grad, x_model.parameters())
            num_parameters = num_parameters + sum([np.prod(p.size()) for p in model_parameters])
    else:
        model_parameters = filter(lambda p: p.requires_grad, model.parameters())
        num_parameters = sum([np.prod(p.size()) for p in model_parameters])
    return num_parameters

def get_param(model):
    paramlist = []
    for name, param in model.named_parameters():
        if param.requires_grad:
            paramlist.append((name, param))
    return paramlist

def init_random_seed(manual_seed):
    """Init random seed."""
    seed = None
    if manual_seed is None:
        seed = random.randint(1, 10000)
    else:
        seed = manual_seed
    print("use random seed: {}".format(seed))
    random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)

def weights_init(m):
    if isinstance(m, gnn.GCNConv):
        init.xavier_normal_(m.weight.data)
        init.constant_(m.bias.data, 0)
    if isinstance(m, gnn.GraphConv):
        init.xavier_normal_(m.weight.data)
    if isinstance(m, gnn.GATConv):
        init.xavier_normal_(m.weight.data)
        init.xavier_normal_(m.att.data)
        init.constant_(m.bias.data, 0)
    if isinstance(m, torch.nn.Linear):
        init.xavier_normal_(m.weight.data)
        init.normal_(m.bias.data)

def init_model(net, restore=None):
    """Init models with cuda and weights."""
    # init weights of model
    # net.apply(weights_init)
    # restore model weights
    if restore is not None and os.path.exists(restore):
        net.load_state_dict(torch.load(restore)['model_state_dict'])
        net.restored = True
        print("Restore model from: {}".format(os.path.abspath(restore)))
    return net

def get_batch_size(input):
    if isinstance(input, list) or isinstance(input, tuple):
        return input[0].size(0)
    if isinstance(input, Data):
        return input.num_graphs
    else:
        return input.size(0)

def read_xlsx(path):
    workbook = load_workbook(path)
    sheet = workbook[workbook.sheetnames[0]]
    data = sheet.values
    cols = next(data)[0:]
    data = list(data)
    data = (itertools.islice(r, 0, None) for r in data)
    df = pd.DataFrame(data, columns = cols)
    return df

class RunningAverage:
    """Computes and stores the average
    """

    def __init__(self):
        self.count = 0
        self.sum = 0
        self.avg = 0

    def update(self, value, n=1):
        self.count += n
        self.sum += value * n
        self.avg = self.sum / self.count